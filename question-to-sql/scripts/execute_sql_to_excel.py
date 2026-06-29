import pandas as pd
import sqlalchemy
from sqlalchemy import create_engine, text
import openpyxl
import json
import sys
import os
from urllib.parse import quote_plus

def get_db_config_from_excel(excel_path):
    try:
        df = pd.read_excel(excel_path, sheet_name='21-数据来源系统')
        if df.empty:
            return None
        
        # 假设配置在第一行有效数据中
        row = df.iloc[0]
        # 根据列名映射
        config = {
            'host': str(row.get('服务器地址', '')),
            'port': str(row.get('端口', '')),
            'dbname': str(row.get('数据库名称', '')),
            'user': str(row.get('数据库用户账号', '')),
            'password': str(row.get('数据库用户密码', ''))
        }
        return config
    except Exception as e:
        print(f"Error reading DB config from Excel: {e}")
        return None

def execute_sql_to_excel(queries_json_path, template_path, output_file='result/测试用例.xlsx'):
    try:
        with open(queries_json_path, 'r', encoding='utf-8') as f:
            queries = json.load(f)
            # 如果 queries 是一个列表，直接使用；如果是字典且包含 'queries' 键，则提取
            if isinstance(queries, dict) and 'queries' in queries:
                queries = queries['queries']
    except Exception as e:
        print(f"Error reading queries file: {e}")
        return

    db_config = get_db_config_from_excel(template_path)
    if not db_config:
        print("Database configuration could not be retrieved from Excel.")
        return

    user = db_config.get('user')
    password = quote_plus(db_config.get('password', ''))
    host = db_config.get('host')
    port = db_config.get('port')
    dbname = db_config.get('dbname')
    
    conn_str = f"mysql+pymysql://{user}:{password}@{host}:{port}/{dbname}"
    
    results = []
    
    try:
        engine = create_engine(conn_str, connect_args={'connect_timeout': 5})
        with engine.connect() as conn:
            for idx, item in enumerate(queries, 1):
                question = item.get('question', '')
                sql = item.get('sql', '')
                
                try:
                    res = conn.execute(text(sql))
                    rows = res.fetchall()
                    if rows:
                        cols = res.keys()
                        formatted_res = []
                        for row in rows:
                            formatted_res.append(", ".join([f"{col}: {val}" for col, val in zip(cols, row)]))
                        result_text = "\n".join(formatted_res)
                    else:
                        result_text = "查询成功，但未匹配到符合条件的数据"
                except Exception as e:
                    result_text = f"SQL执行出错: {str(e)}"
                
                results.append({
                    "序号": idx,
                    "问句": question,
                    "执行SQL": sql,
                    "执行结果": result_text
                })
    except Exception as e:
        for idx, item in enumerate(queries, 1):
            results.append({
                "序号": idx,
                "问句": item.get('question', ''),
                "执行SQL": item.get('sql', ''),
                "执行结果": f"数据库连接失败: {str(e)}"
            })

    df = pd.DataFrame(results)
    
    # 样式美化
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='测试用例')
        workbook = writer.book
        worksheet = writer.sheets['测试用例']
        
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # 表头样式
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        header_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            
        # 数据单元格样式
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))
        top_alignment = Alignment(vertical='top', wrap_text=True)
        
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = top_alignment
        
        # 列宽设置
        column_widths = {'A': 8, 'B': 40, 'C': 60, 'D': 50}
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

    print(f"Results saved to {output_file}")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python execute_sql_to_excel.py <queries_json_path> <template_excel_path>")
    else:
        execute_sql_to_excel(sys.argv[1], sys.argv[2])
